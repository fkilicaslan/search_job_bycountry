#!/usr/bin/env python3
"""Generate Excel with all scanned jobs from pipeline.md and applications.md.

German language scoring rules applied on top of batch evaluator scores:
  EXCLUDE tier  -> cap score at 1.5 (only when not already flagged in notes)
  REDUCE tier   -> deduct 0.4 from score, add obstacle note
"""

import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

PIPELINE     = Path("career-ops/data/pipeline.md")
APPLICATIONS = Path("career-ops/data/applications.md")
REPORTS_DIR  = Path("career-ops/reports")
OUTPUT       = Path("career-ops/data/all-jobs.xlsx")

# ── German language detection ──────────────────────────────────────────────────

# If existing notes already mention German as a blocker/risk, skip re-evaluation
_NOTES_FLAGGED = re.compile(
    r"(?:german|deutsch).{0,50}(?:blocker|dealbreaker|hard|native|c2|gap|risk)"
    r"|(?:blocker|dealbreaker).{0,50}(?:german|deutsch)"
    r"|c1 vs|verhandlungs|muttersprachl"
    r"|german.*fluency required|fluent german.*required",
    re.IGNORECASE,
)

# EXCLUDE tier: native / verhandlungssicher / C2 -> score cap
_EXCLUDE = re.compile(
    r"native german|german native|native speaker.{0,20}german|german.{0,20}native speaker"
    r"|muttersprachlich|verhandlungssicher|deutsch c2\b|c2 german"
    r"|muttersprachenniveau"
    r"|hard dealbreaker.*german|german.*hard dealbreaker"
    r"|must have fluent german|fluent german required",
    re.IGNORECASE,
)

# REDUCE tier: business-fluent / fliesend / excellent German in client context
_REDUCE = re.compile(
    r"business.{1,10}fluent.{1,15}german|business.{1,10}level.{1,15}german"
    r"|flie.end.{1,20}deutsch|deutsch.{1,20}flie.end"
    r"|flie.end in wort"
    r"|sehr gut.{1,20}deutsch|ausgezeichnet.{1,20}deutsch"
    r"|excellent.{1,15}german.{1,15}(?:skill|communicat|negotiat|present)"
    r"|strong.{1,15}german.{1,15}(?:skill|communicat|present)"
    r"|negotiation.{1,40}presentation.{1,40}german"
    r"|communication.{1,40}negotiation.{1,40}german"
    r"|german.{1,40}(?:communication|negotiation|presentation).{1,40}skill",
    re.IGNORECASE,
)

# Safe patterns override REDUCE (German is nice-to-have)
_SAFE = re.compile(
    r"german.{1,25}(?:plus|nice.to.have|preferred|advantage|bonus)"
    r"|(?:plus|advantage|bonus|prefer).{1,25}german"
    r"|von vorteil|german c1 or above|conversational german",
    re.IGNORECASE,
)

REDUCE_NOTE = (
    "German language gap: role requires business-fluent German, candidate is C1 "
    "(sufficient for written/internal use, not for native-level client sales conversations)"
)
EXCLUDE_NOTE = (
    "German hard blocker: role requires native/verhandlungssicher/C2-level German "
    "-- C1 not sufficient"
)

REDUCE_DEDUCTION = 0.40
EXCLUDE_SCORE_CAP = 1.50


def _load_report(report_col_value):
    """Read report file from '[N](reports/name.md)' cell value. Returns text or None."""
    if not report_col_value:
        return None
    m = re.search(r'\(reports/(.+?)\)', report_col_value)
    if not m:
        return None
    path = REPORTS_DIR / m.group(1)
    if not path.exists():
        return None
    try:
        return path.read_text(encoding="utf-8", errors="replace")
    except Exception:
        return None


def detect_german_tier(report_col_value, existing_notes):
    """Return 'exclude', 'reduce', or None. Skip if notes already flag German."""
    if _NOTES_FLAGGED.search(existing_notes or ""):
        return None
    content = _load_report(report_col_value)
    if not content:
        return None
    # Safe context overrides REDUCE but not EXCLUDE
    if _SAFE.search(content):
        return "exclude" if _EXCLUDE.search(content) else None
    if _EXCLUDE.search(content):
        return "exclude"
    if _REDUCE.search(content):
        return "reduce"
    return None


# ── Parse applications.md ─────────────────────────────────────────────────────

SCORE_RE = re.compile(r"^\d+\.\d+/5$")
CANONICAL_STATUSES = {
    "Evaluated", "Applied", "Responded", "Interview", "Offer",
    "Rejected", "Discarded", "SKIP", "N/A",
}


def parse_applications():
    rows = []
    with open(APPLICATIONS, encoding="utf-8") as f:
        for line in f:
            line = line.rstrip()
            if not line.startswith("|"):
                continue
            parts = [p.strip() for p in line.split("|")]
            parts = [p for p in parts if p != ""]
            if len(parts) < 8:
                continue
            if parts[0] == "#":
                continue
            if all(set(p) <= set("-") for p in parts):
                continue
            if not parts[0].isdigit():
                continue
            try:
                num     = parts[0]
                date    = parts[1]
                company = parts[2]

                score_str = ""
                status    = ""
                score_idx = None
                for i in range(3, min(len(parts), 9)):
                    if SCORE_RE.match(parts[i]) or parts[i] == "N/A":
                        score_idx = i
                        score_str = parts[i]
                        break
                if score_idx is None:
                    for i in range(3, min(len(parts), 8)):
                        if parts[i] in CANONICAL_STATUSES:
                            score_idx = i
                            score_str = ""
                            status    = parts[i]
                            break
                if score_idx is not None:
                    role       = " | ".join(parts[3:score_idx])
                    if not status:
                        status = parts[score_idx + 1] if score_idx + 1 < len(parts) else ""
                    report_col = parts[score_idx + 3] if score_idx + 3 < len(parts) else ""
                    notes      = parts[score_idx + 4] if score_idx + 4 < len(parts) else ""
                else:
                    role       = parts[3] if len(parts) > 3 else ""
                    report_col = ""
                    notes      = ""

                if status not in CANONICAL_STATUSES:
                    status = "Discarded"

                rows.append({
                    "num":        num,
                    "date":       date,
                    "company":    company,
                    "role":       role,
                    "score_str":  score_str,
                    "status":     status,
                    "report_col": report_col,
                    "notes":      notes,
                    "url":        "",
                })
            except Exception:
                continue
    return rows


# ── Parse pipeline.md ─────────────────────────────────────────────────────────

def parse_pipeline():
    pending, evaluated = [], []
    with open(PIPELINE, encoding="utf-8") as f:
        for line in f:
            line = line.rstrip()
            if not line.startswith("- "):
                continue
            m = re.match(
                r"- \[x\] #(\d+) \| (https?://\S+) \| (.+?) \| (.+?)(?:\s+\|\s+(.+?))?$",
                line,
            )
            if m:
                evaluated.append({
                    "num":     m.group(1),
                    "url":     m.group(2),
                    "company": m.group(3).strip(),
                    "role":    m.group(4).strip(),
                })
                continue
            m = re.match(r"- \[ \] (https?://\S+) \| (.+?) \| (.+)$", line)
            if m:
                pending.append({
                    "url":     m.group(1),
                    "company": m.group(2).strip(),
                    "role":    m.group(3).strip(),
                })
                continue
            m = re.match(r"- \[ \] (https?://\S+)$", line)
            if m:
                pending.append({"url": m.group(1), "company": "", "role": ""})
    return pending, evaluated


# ── Build combined dataset ────────────────────────────────────────────────────

def to_numeric_score(s):
    if not s:
        return None
    m = re.match(r"^(\d+\.\d+)/5$", s)
    return float(m.group(1)) if m else None


def build_dataset():
    app_rows         = parse_applications()
    pending, evaluated = parse_pipeline()

    app_by_num = {r["num"]: r for r in app_rows}
    for ev in evaluated:
        if ev["num"] in app_by_num:
            app_by_num[ev["num"]]["url"] = ev["url"]

    final            = []
    german_reduce_ct = 0
    german_exclude_ct = 0

    for r in app_rows:
        base_score = to_numeric_score(r["score_str"])
        notes      = r["notes"]
        report_col = r["report_col"]

        german_tier = detect_german_tier(report_col, notes)
        score    = base_score
        obstacle = notes

        if german_tier == "reduce" and base_score is not None and base_score > 1.0:
            score    = round(max(base_score - REDUCE_DEDUCTION, 1.0), 2)
            obstacle = f"{REDUCE_NOTE} | {notes}" if notes else REDUCE_NOTE
            german_reduce_ct += 1
        elif german_tier == "exclude" and base_score is not None and base_score > EXCLUDE_SCORE_CAP:
            score    = EXCLUDE_SCORE_CAP
            obstacle = f"{EXCLUDE_NOTE} | {notes}" if notes else EXCLUDE_NOTE
            german_exclude_ct += 1

        final.append({
            "source":     "Evaluated",
            "company":    r["company"],
            "role":       r["role"],
            "url":        r["url"],
            "score":      score,
            "status":     r["status"],
            "bottleneck": obstacle,
        })

    for p in pending:
        final.append({
            "source":     "Pending",
            "company":    p["company"],
            "role":       p["role"],
            "url":        p["url"],
            "score":      None,
            "status":     "Pending",
            "bottleneck": "",
        })

    print(f"German REDUCE adjustments (-{REDUCE_DEDUCTION}): {german_reduce_ct}")
    print(f"German EXCLUDE adjustments (cap {EXCLUDE_SCORE_CAP}): {german_exclude_ct}")
    return final


# ── Write Excel ───────────────────────────────────────────────────────────────

STATUS_COLORS = {
    "Applied":   "C6EFCE",
    "Interview": "C6EFCE",
    "Offer":     "FFEB9C",
    "Evaluated": "DDEBF7",
    "Pending":   "F2F2F2",
    "Discarded": "F4CCCC",
    "SKIP":      "F4CCCC",
    "Rejected":  "F4CCCC",
}


def col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width


def write_excel(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Jobs"

    headers = ["#", "Company", "Position", "Score", "Status", "URL", "Bottleneck / Notes"]
    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="2F3640")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 22

    thin   = Side(border_style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, r in enumerate(rows, 2):
        status     = r["status"]
        fill_color = STATUS_COLORS.get(status, "FFFFFF")
        row_fill   = PatternFill("solid", fgColor=fill_color)

        cells = [
            (1, str(i - 1)),
            (2, r["company"]),
            (3, r["role"]),
            (4, r["score"]),
            (5, r["status"]),
            (6, r["url"]),
            (7, r["bottleneck"]),
        ]
        for col, val in cells:
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill      = row_fill
            cell.border    = border
            cell.alignment = Alignment(vertical="top", wrap_text=(col in (3, 7)))
            if col == 4:
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="center", vertical="top")
            if col == 5:
                cell.alignment = Alignment(horizontal="center", vertical="top")

        url_cell = ws.cell(row=i, column=6)
        if r["url"]:
            url_cell.hyperlink = r["url"]
            url_cell.font      = Font(color="1155CC", underline="single")
            url_cell.value     = r["url"]

    col_width(ws, "A",  6)
    col_width(ws, "B", 26)
    col_width(ws, "C", 42)
    col_width(ws, "D", 10)
    col_width(ws, "E", 12)
    col_width(ws, "F", 60)
    col_width(ws, "G", 70)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{len(rows)+1}"

    wb.save(OUTPUT)
    print(f"Saved {len(rows)} rows -> {OUTPUT}")


# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    rows = build_dataset()
    write_excel(rows)
    by_status = {}
    for r in rows:
        by_status[r["status"]] = by_status.get(r["status"], 0) + 1
    print("Status breakdown:")
    for s, n in sorted(by_status.items(), key=lambda x: -x[1]):
        print(f"  {s}: {n}")
